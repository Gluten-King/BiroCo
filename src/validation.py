#!/usr/bin/env python3
"""Validate generated BiroCo 3NF workbook and output audit workbook."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from itertools import groupby
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Directory containing this script; all default paths are resolved relative to it.
_HERE = Path(__file__).parent


def _res(p: str) -> Path:
    """Return p as a Path, resolved relative to _HERE if not absolute."""
    path = Path(p)
    return path if path.is_absolute() else _HERE / path


DEFAULT_INPUT_XLSX = "biroco_3nf_generated.xlsx"
DEFAULT_REPORT_XLSX = "biroco_validation_log.xlsx"
NULL = "NULL"

REQUIRED_SHEETS = [
    "OrderTable", "Customer", "Product", "Orderline", "Supplier", "SupplierProduct",
    "Discount", "Inventory", "Delivery", "Return", "FlatView",
]

ID_PATTERNS = {
    "OrderID": r"^O\d{5}$",
    "CustomerID": r"^C\d{5}$",
    "ProductID": r"^P\d{5}$",
    "OrderlineID": r"^OL\d{5}$",
    "SupplierID": r"^S\d{5}$",
    "DiscountID": r"^D\d{5}$",
    "InventoryID": r"^I00[ABC]\d+$",
    "DeliveryID": r"^DLV\d{5}$",
    "ReturnID": r"^R\d{5}$",
}

FULFILLMENT_ALLOWED = {"pending", "packed", "shipped", "partially_delivered", "delivered", "cancelled"}
AFTERSALES_ALLOWED = {"no_return", "return_in_progress", "refunded", "return_rejected", "partially_refunded"}
DELIVERY_ALLOWED = {"pending", "packed", "shipped", "delivered"}
RETURN_ALLOWED = {"requested", "approved", "rejected", "refund"}
PLATFORM_ALLOWED = {
    "website", "app", "marketplace_amazon", "marketplace_temu",
    "marketplace_ebay", "marketplace_etsy", "marketplace_walmart",
}
GB_PHONE = re.compile(r"^07\d{9}$")
MONEY_Q = Decimal("0.01")
GROSS_MARGIN_MIN = Decimal("0.30")
GROSS_MARGIN_MAX = Decimal("0.70")
TRACKING_PATTERNS = [
    re.compile(r"^[A-Z]{2}\d{9}GB$"),       # Royal Mail-like
    re.compile(r"^H00AA\d{11}$"),           # Evri-like
    re.compile(r"^\d{10}$"),                # DHL-like
    re.compile(r"^1Z[0-9A-Z]{16}$"),        # UPS-like
]

# Distribution targets
TARGET_DELIVERED_RATIO = Decimal("0.65")
DELIVERED_RATIO_MIN = Decimal("0.60")
DELIVERED_RATIO_MAX = Decimal("0.72")

TARGET_COMMENT_RATIO_COMPLETED = Decimal("0.50")
COMMENT_RATIO_MIN = Decimal("0.42")
COMMENT_RATIO_MAX = Decimal("0.58")

TARGET_RETURN_RATE = Decimal("0.10")
RETURN_RATE_MIN = Decimal("0.06")
RETURN_RATE_MAX = Decimal("0.14")

TARGET_SCORE_BANDS = {
    "0":   Decimal("0.01"),
    "1-3": Decimal("0.04"),
    "4-6": Decimal("0.10"),
    "7-8": Decimal("0.55"),
    "9-10": Decimal("0.30"),
}
SCORE_BAND_TOL = Decimal("0.10")

# ID gap policy
ID_GAP_MAX_RATE = Decimal("0.03")
ID_GAP_MIN_SIZE = 5
ID_GAP_MAX_SIZE = 30
NO_GAP_ID_KEYS = {"SupplierID", "ProductID", "InventoryID"}
GAP_CHECK_ID_KEYS = {"OrderID", "CustomerID", "OrderlineID", "DeliveryID", "ReturnID"}


@dataclass
class CheckResult:
    check_id: str
    category: str
    status: str
    details: str


class Collector:
    def __init__(self):
        self.rows = []

    def add(self, cid, cat, ok, ok_msg, fail_msg):
        self.rows.append(CheckResult(cid, cat, "PASS" if ok else "FAIL", ok_msg if ok else fail_msg))

    def warn(self, cid, cat, details):
        self.rows.append(CheckResult(cid, cat, "WARN", details))


def n(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.upper() == NULL:
        return None
    return s


def d(v):
    s = n(v)
    if s is None:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def q_money(v: Decimal) -> Decimal:
    return v.quantize(MONEY_Q, rounding=ROUND_HALF_UP)


def clamp_decimal(v: Decimal, low: Decimal, high: Decimal) -> Decimal:
    if v < low:
        return low
    if v > high:
        return high
    return v


def dt(v) -> date | None:
    s = n(v)
    if s is None:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def dec_ratio(num: int, den: int) -> Decimal:
    if den <= 0:
        return Decimal("0")
    return Decimal(num) / Decimal(den)


def score_band(score: int) -> str:
    if score == 0:
        return "0"
    if 1 <= score <= 3:
        return "1-3"
    if 4 <= score <= 6:
        return "4-6"
    if 7 <= score <= 8:
        return "7-8"
    return "9-10"


def id_gap_stats(values: set[str]) -> tuple[int, int, Decimal, int, int]:
    nums = []
    for v in values:
        if v is None:
            continue
        m = re.search(r"(\d+)$", v)
        if m is not None:
            nums.append(int(m.group(1)))
    nums = sorted(set(nums))
    if len(nums) <= 1:
        return 0, 0, Decimal("0"), 0, 0

    events = 0
    bad_size = 0
    max_gap = 0
    total_missing = 0
    for i in range(1, len(nums)):
        diff = nums[i] - nums[i - 1]
        if diff > 1:
            events += 1
            gap_size = diff - 1
            total_missing += gap_size
            if gap_size < ID_GAP_MIN_SIZE or gap_size > ID_GAP_MAX_SIZE:
                bad_size += 1
            if gap_size > max_gap:
                max_gap = gap_size
    rate = dec_ratio(events, len(nums) - 1)
    return events, bad_size, rate, max_gap, total_missing


def load_tables(path: Path):
    wb = load_workbook(path, data_only=True)
    out = {}
    for s in REQUIRED_SHEETS:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[s] = []
            continue
        hdr = [str(x).strip() if x is not None else "" for x in rows[0]]
        data = []
        for r in rows[1:]:
            item = {hdr[i]: (r[i] if i < len(r) else None) for i in range(len(hdr)) if hdr[i]}
            if any(v is not None and str(v).strip() != "" for v in item.values()):
                data.append(item)
        out[s] = data
    return out


def pk_unique(c: Collector, table, rows, keys):
    seen, blank, dup = set(), 0, 0
    for row in rows:
        k = tuple(n(row.get(x)) for x in keys)
        if any(x is None for x in k):
            blank += 1
            continue
        if k in seen:
            dup += 1
        else:
            seen.add(k)
    c.add(f"PK_{table}", "PK", blank == 0 and dup == 0, f"{table} pk valid", f"{table} blank={blank} dup={dup}")


def fk(c: Collector, cid, rows, col, parent, nullable=False):
    bad = []
    for row in rows:
        v = n(row.get(col))
        if v is None:
            if not nullable:
                bad.append("<NULL>")
            continue
        if v not in parent:
            bad.append(v)
    u = sorted(set(bad))
    c.add(cid, "FK", len(u) == 0, f"{cid} valid", f"{cid} missing={len(u)} sample={u[:5]}")


def pattern(c: Collector, name, values, pat):
    bad = sorted({x for x in values if x is not None and re.match(pat, x) is None})
    c.add(f"ID_{name}", "FORMAT", len(bad) == 0, f"{name} format valid", f"{name} invalid={len(bad)} sample={bad[:5]}")


def allowed(c: Collector, cid, values, options):
    bad = sorted({x for x in values if x is not None and x not in options})
    c.add(cid, "ENUM", len(bad) == 0, f"{cid} valid", f"{cid} invalid={len(bad)} sample={bad[:5]}")


def derive_fulfillment(cancelled: bool, d_statuses):
    if cancelled:
        return "cancelled"
    if not d_statuses:
        return "pending"
    s = set(d_statuses)
    if s == {"pending"}:
        return "pending"
    if s == {"packed"}:
        return "packed"
    if "delivered" in s:
        return "delivered" if s == {"delivered"} else "partially_delivered"
    if "shipped" in s:
        return "shipped"
    if "packed" in s:
        return "packed"
    return "pending"


def derive_aftersales(ret_statuses):
    if not ret_statuses:
        return "no_return"
    s = set(ret_statuses)
    if "requested" in s or "approved" in s:
        return "return_in_progress"
    if s == {"refund"}:
        return "refunded"
    if s == {"rejected"}:
        return "return_rejected"
    if s.issubset({"refund", "rejected"}):
        return "partially_refunded"
    return "return_in_progress"

def validate(c: Collector, t):
    orders = t["OrderTable"]; customers = t["Customer"]; products = t["Product"]; orderlines = t["Orderline"]
    suppliers = t["Supplier"]; supplier_products = t["SupplierProduct"]; discounts = t["Discount"]
    inventory = t["Inventory"]; deliveries = t["Delivery"]; returns = t["Return"]; flat = t["FlatView"]

    ocols = set(orders[0].keys()) if orders else set()
    ccols = set(customers[0].keys()) if customers else set()
    pcols = set(products[0].keys()) if products else set()
    olcols = set(orderlines[0].keys()) if orderlines else set()
    scols = set(suppliers[0].keys()) if suppliers else set()
    spcols = set(supplier_products[0].keys()) if supplier_products else set()
    discols = set(discounts[0].keys()) if discounts else set()
    icols = set(inventory[0].keys()) if inventory else set()
    dcols = set(deliveries[0].keys()) if deliveries else set()
    rcols = set(returns[0].keys()) if returns else set()
    fcols = set(flat[0].keys()) if flat else set()

    c.add("SCHEMA_Order", "SCHEMA", {"FulfillmentStatus", "AfterSalesStatus"}.issubset(ocols) and "OrderStatus" not in ocols,
          "order schema ok", f"order cols={sorted(ocols)}")
    c.add("SCHEMA_Orderline", "SCHEMA", "UnitPriceAtPurchase" not in olcols and {"DeliveryID", "ReturnID"}.issubset(olcols),
          "orderline schema ok", f"orderline cols={sorted(olcols)}")
    c.add("SCHEMA_SupplierProduct", "SCHEMA", {"LeadTimeDays", "UnitCostGBP"}.issubset(spcols),
          "supplierproduct schema ok", f"supplierproduct cols={sorted(spcols)}")
    c.add("SCHEMA_Delivery", "SCHEMA", "OrderID" not in dcols, "delivery schema ok", f"delivery cols={sorted(dcols)}")
    c.add("SCHEMA_Return", "SCHEMA", "OrderlineID" not in rcols and "DeliveryID" not in rcols,
          "return schema ok", f"return cols={sorted(rcols)}")
    c.add("SCHEMA_FlatKeys", "SCHEMA", {"OrderID", "CustomerID", "ProductID", "OrderlineID", "SupplierID", "DiscountID", "DeliveryID", "ReturnID", "InventoryID"}.issubset(fcols),
          "flat key cols present", f"flat cols={sorted(fcols)}")
    nf_union_cols = ocols | ccols | pcols | olcols | scols | spcols | discols | icols | dcols | rcols
    flat_missing_cols = sorted(nf_union_cols - fcols)
    c.add(
        "SCHEMA_FlatAllNFCols",
        "SCHEMA",
        len(flat_missing_cols) == 0,
        "flat contains all NF columns",
        f"missing flat cols={flat_missing_cols}",
    )

    empty = 0; bad_null = 0
    for rows in t.values():
        for row in rows:
            for v in row.values():
                if v is None:
                    continue
                s = str(v)
                if s == "":
                    empty += 1
                if s.lower() == "null" and s != "NULL":
                    bad_null += 1
    c.add("NULL_EMPTY", "QUALITY", empty == 0, "no empty strings", f"empty strings={empty}")
    c.add("NULL_UPPER", "QUALITY", bad_null == 0, "uppercase NULL only", f"bad null cells={bad_null}")

    pk_unique(c, "OrderTable", orders, ["OrderID"])
    pk_unique(c, "Customer", customers, ["CustomerID"])
    pk_unique(c, "Product", products, ["ProductID"])
    pk_unique(c, "Orderline", orderlines, ["OrderlineID"])
    pk_unique(c, "Supplier", suppliers, ["SupplierID"])
    pk_unique(c, "SupplierProduct", supplier_products, ["SupplierID", "ProductID"])
    pk_unique(c, "Discount", discounts, ["DiscountID"])
    pk_unique(c, "Inventory", inventory, ["InventoryID"])
    pk_unique(c, "Delivery", deliveries, ["DeliveryID"])
    pk_unique(c, "Return", returns, ["ReturnID"])

    order_ids = {n(x.get("OrderID")) for x in orders}; customer_ids = {n(x.get("CustomerID")) for x in customers}
    product_ids = {n(x.get("ProductID")) for x in products}; orderline_ids = {n(x.get("OrderlineID")) for x in orderlines}
    supplier_ids = {n(x.get("SupplierID")) for x in suppliers}; discount_ids = {n(x.get("DiscountID")) for x in discounts}
    inventory_ids = {n(x.get("InventoryID")) for x in inventory}; delivery_ids = {n(x.get("DeliveryID")) for x in deliveries}; return_ids = {n(x.get("ReturnID")) for x in returns}

    for k, vals in [("OrderID", order_ids), ("CustomerID", customer_ids), ("ProductID", product_ids), ("OrderlineID", orderline_ids),
                    ("SupplierID", supplier_ids), ("DiscountID", discount_ids), ("InventoryID", inventory_ids), ("DeliveryID", delivery_ids), ("ReturnID", return_ids)]:
        pattern(c, k, vals, ID_PATTERNS[k])

    fk(c, "FK_Order_Customer", orders, "CustomerID", {x for x in customer_ids if x})
    fk(c, "FK_Order_Discount", orders, "DiscountID", {x for x in discount_ids if x})
    fk(c, "FK_Orderline_Order", orderlines, "OrderID", {x for x in order_ids if x})
    fk(c, "FK_Orderline_Product", orderlines, "ProductID", {x for x in product_ids if x})
    fk(c, "FK_Orderline_Supplier", orderlines, "SupplierID", {x for x in supplier_ids if x})
    fk(c, "FK_Orderline_Delivery", orderlines, "DeliveryID", {x for x in delivery_ids if x}, nullable=True)
    fk(c, "FK_Orderline_Return", orderlines, "ReturnID", {x for x in return_ids if x}, nullable=True)
    fk(c, "FK_Inventory_Product", inventory, "ProductID", {x for x in product_ids if x})
    fk(c, "FK_SP_Supplier", supplier_products, "SupplierID", {x for x in supplier_ids if x})
    fk(c, "FK_SP_Product", supplier_products, "ProductID", {x for x in product_ids if x})
    sp_pairs = {
        (n(r.get("SupplierID")), n(r.get("ProductID")))
        for r in supplier_products
        if n(r.get("SupplierID")) and n(r.get("ProductID"))
    }
    missing_ol_pair = []
    for ol in orderlines:
        pair = (n(ol.get("SupplierID")), n(ol.get("ProductID")))
        if pair not in sp_pairs:
            missing_ol_pair.append((n(ol.get("OrderlineID")), pair[0], pair[1]))
    c.add(
        "FK_Orderline_SupplierProductPair",
        "FK",
        len(missing_ol_pair) == 0,
        f"all orderline (SupplierID,ProductID) pairs valid; checked={len(orderlines)}",
        f"missing pairs={len(missing_ol_pair)} sample={missing_ol_pair[:5]}",
    )

    sp_rows_by_product = defaultdict(list)
    sp_cost_by_pair = {}
    sp_bad_leadtime = 0
    sp_bad_cost = 0
    margin_out_of_range = 0
    margin_values = []
    for sp in supplier_products:
        sid = n(sp.get("SupplierID"))
        pid = n(sp.get("ProductID"))
        lead_raw = n(sp.get("LeadTimeDays"))
        cost = d(sp.get("UnitCostGBP"))
        lead = None
        if lead_raw is None:
            sp_bad_leadtime += 1
        else:
            try:
                lead = int(lead_raw)
                if lead <= 0:
                    sp_bad_leadtime += 1
                    lead = None
            except ValueError:
                sp_bad_leadtime += 1
        if cost is None or cost <= 0:
            sp_bad_cost += 1
        if sid and pid:
            sp_cost_by_pair[(sid, pid)] = cost
        if pid and lead is not None and cost is not None and cost > 0:
            sp_rows_by_product[pid].append((lead, cost))

    product_price_by_id = {n(p.get("ProductID")): d(p.get("UnitPriceGBP")) for p in products}
    for pid, rows in sp_rows_by_product.items():
        unit_price = product_price_by_id.get(pid)
        if unit_price is None or unit_price <= 0:
            continue
        for _lead, cost in rows:
            gross_margin = (unit_price - cost) / unit_price
            margin_values.append(gross_margin)
            if gross_margin < GROSS_MARGIN_MIN or gross_margin > GROSS_MARGIN_MAX:
                margin_out_of_range += 1

    trend_pairs = 0
    trend_inversions = 0
    for _pid, rows in sp_rows_by_product.items():
        for i in range(len(rows)):
            for j in range(i + 1, len(rows)):
                l1, c1 = rows[i]
                l2, c2 = rows[j]
                if l1 == l2:
                    continue
                trend_pairs += 1
                if l1 > l2 and c1 > c2:
                    trend_inversions += 1
                if l2 > l1 and c2 > c1:
                    trend_inversions += 1

    c.add("RULE_SP_LeadTimePositive", "RULE", sp_bad_leadtime == 0, "LeadTimeDays positive int", f"bad LeadTimeDays rows={sp_bad_leadtime}")
    c.add("RULE_SP_UnitCostPositive", "RULE", sp_bad_cost == 0, "UnitCostGBP positive", f"bad UnitCostGBP rows={sp_bad_cost}")
    c.add(
        "RULE_SP_GrossMarginRange",
        "RULE",
        margin_out_of_range == 0,
        f"all supplier-product gross margins in [{GROSS_MARGIN_MIN:.0%}, {GROSS_MARGIN_MAX:.0%}]",
        f"gross margin out-of-range rows={margin_out_of_range}",
    )
    c.add(
        "RULE_SP_LeadTimeCostTrend",
        "RULE",
        trend_inversions == 0,
        "longer lead time does not increase procurement cost on comparable pairs",
        f"trend inversions={trend_inversions}, comparable_pairs={trend_pairs}",
    )

    allowed(c, "ENUM_Fulfillment", (n(x.get("FulfillmentStatus")) for x in orders), FULFILLMENT_ALLOWED)
    allowed(c, "ENUM_AfterSales", (n(x.get("AfterSalesStatus")) for x in orders), AFTERSALES_ALLOWED)
    allowed(c, "ENUM_Delivery", (n(x.get("DeliveryStatus")) for x in deliveries), DELIVERY_ALLOWED)
    allowed(c, "ENUM_Return", (n(x.get("ReturnStatus")) for x in returns), RETURN_ALLOWED)
    allowed(c, "ENUM_Platform", (n(x.get("Platform")) for x in orders), PLATFORM_ALLOWED)

    bad_cp = sum(1 for x in customers if GB_PHONE.match(n(x.get("CustomerPhone")) or "") is None)
    bad_sp = sum(1 for x in suppliers if GB_PHONE.match(n(x.get("SupplierPhone")) or "") is None)
    c.add("FORMAT_CustomerPhone", "FORMAT", bad_cp == 0, "customer phone format ok", f"bad customer phones={bad_cp}")
    c.add("FORMAT_SupplierPhone", "FORMAT", bad_sp == 0, "supplier phone format ok", f"bad supplier phones={bad_sp}")

    inv_by_id = {n(x.get("InventoryID")): x for x in inventory}
    bad_inv_loc = 0
    for iid, row in inv_by_id.items():
        loc = n(row.get("Location"))
        m = re.match(r"^I00([ABC])\d+$", iid or "")
        if m is None or loc not in {"WarehouseA", "WarehouseB", "WarehouseC"} or m.group(1) != loc[-1]:
            bad_inv_loc += 1
    c.add("RULE_InventoryID_Location", "RULE", bad_inv_loc == 0, "inventory id matches location", f"mismatch={bad_inv_loc}")

    order_by_id = {n(x.get("OrderID")): x for x in orders}; delivery_by_id = {n(x.get("DeliveryID")): x for x in deliveries}; return_by_id = {n(x.get("ReturnID")): x for x in returns}
    ol_by_id = {n(x.get("OrderlineID")): x for x in orderlines}
    ol_by_order = defaultdict(list); ol_by_delivery = defaultdict(list); ol_by_return = defaultdict(list)
    for ol in orderlines:
        oid = n(ol.get("OrderID")); did = n(ol.get("DeliveryID")); rid = n(ol.get("ReturnID"))
        if oid: ol_by_order[oid].append(ol)
        if did: ol_by_delivery[did].append(ol)
        if rid: ol_by_return[rid].append(ol)

    d_conflict = 0; d_empty = 0
    for did in [x for x in delivery_ids if x]:
        links = ol_by_delivery.get(did, [])
        if not links:
            d_empty += 1; continue
        if len({n(x.get("OrderID")) for x in links}) != 1:
            d_conflict += 1
    c.add("CONS_Delivery_OneOrder", "CONSISTENCY", d_conflict == 0, "delivery->order unique", f"conflict={d_conflict}")
    c.add("CONS_Delivery_HasOrderline", "CONSISTENCY", d_empty == 0, "delivery has orderline", f"delivery without orderline={d_empty}")

    # Strict warehouse match: every product in a Delivery's Orderlines must have
    # an inventory record at that Delivery's Warehouse (shipping wh == storage wh).
    inv_locs_by_product: dict = defaultdict(set)
    for inv_row in inventory:
        pid_v = n(inv_row.get("ProductID"))
        loc_v = n(inv_row.get("Location"))
        if pid_v and loc_v:
            inv_locs_by_product[pid_v].add(loc_v)

    bad_del_wh = 0
    bad_del_wh_sample = []
    checked_del_wh = 0
    for did_v in [x for x in delivery_ids if x]:
        wh = n(delivery_by_id[did_v].get("Warehouse"))
        if wh is None:
            continue
        lines = ol_by_delivery.get(did_v, [])
        if not lines:
            continue
        checked_del_wh += 1
        # Each product in this delivery must have inventory at the delivery warehouse.
        mismatch_pids = []
        for ol in lines:
            pid_v = n(ol.get("ProductID"))
            if pid_v:
                valid_whs_for_pid = inv_locs_by_product.get(pid_v, set())
                if valid_whs_for_pid and wh not in valid_whs_for_pid:
                    mismatch_pids.append(pid_v)
        if mismatch_pids:
            bad_del_wh += 1
            if len(bad_del_wh_sample) < 5:
                bad_del_wh_sample.append((did_v, wh, sorted(set(mismatch_pids))))
    c.add(
        "RULE_Delivery_WarehouseMatchesInventory",
        "RULE",
        bad_del_wh == 0,
        f"all products in delivery are stocked at delivery warehouse (intersection); checked={checked_del_wh}",
        f"warehouse/inventory mismatch={bad_del_wh} sample={bad_del_wh_sample}",
    )

    cancel_with_delivery = 0; noncancel_without_delivery = 0; fulfill_mismatch = 0
    for oid, o in order_by_id.items():
        fs = n(o.get("FulfillmentStatus")); lines = ol_by_order.get(oid, [])
        dids = [n(x.get("DeliveryID")) for x in lines if n(x.get("DeliveryID"))]
        d_statuses = [n(delivery_by_id[d].get("DeliveryStatus")) for d in dids if d in delivery_by_id]
        if fs == "cancelled" and dids: cancel_with_delivery += 1
        if fs != "cancelled" and not dids: noncancel_without_delivery += 1
        if derive_fulfillment(fs == "cancelled", [x for x in d_statuses if x]) != fs: fulfill_mismatch += 1
    c.add("LIFE_Cancelled_NoDelivery", "LIFECYCLE", cancel_with_delivery == 0, "cancelled has no delivery", f"violations={cancel_with_delivery}")
    c.add("LIFE_NonCancelled_HasDelivery", "LIFECYCLE", noncancel_without_delivery == 0, "non-cancelled has delivery", f"violations={noncancel_without_delivery}")
    c.add("RULE_FulfillmentDerived", "RULE", fulfill_mismatch == 0, "fulfillment derived ok", f"mismatches={fulfill_mismatch}")

    ret_no_ol = 0; ret_qty_bad = 0; ret_non_delivered = 0; ret_multi_order = 0
    for rid, rr in return_by_id.items():
        links = ol_by_return.get(rid, [])
        if not links:
            ret_no_ol += 1; continue
        qty_sum = Decimal("0"); oids = set()
        for ol in links:
            q = d(ol.get("Quantity"))
            if q is not None: qty_sum += q
            oids.add(n(ol.get("OrderID")))
            did = n(ol.get("DeliveryID")); ds = n(delivery_by_id.get(did, {}).get("DeliveryStatus")) if did else None
            if ds != "delivered": ret_non_delivered += 1
        rq = d(rr.get("ReturnQty"))
        if rq is None or rq != qty_sum: ret_qty_bad += 1
        if len(oids) > 1: ret_multi_order += 1
    c.add("CARD_Return_HasOrderline", "CARDINALITY", ret_no_ol == 0, "return linked orderline exists", f"return without orderline={ret_no_ol}")
    c.add("CONS_ReturnQty", "CONSISTENCY", ret_qty_bad == 0, "return qty sum matches", f"qty mismatch={ret_qty_bad}")
    c.add("LIFE_Return_OnDelivered", "LIFECYCLE", ret_non_delivered == 0, "returns on delivered only", f"violations={ret_non_delivered}")
    c.add("CONS_Return_OneOrder", "CONSISTENCY", ret_multi_order == 0, "return lines from one order", f"multi-order returns={ret_multi_order}")

    after_bad = 0
    for oid, o in order_by_id.items():
        fs = n(o.get("FulfillmentStatus")); actual = n(o.get("AfterSalesStatus"))
        rs = []
        for ol in ol_by_order.get(oid, []):
            rid = n(ol.get("ReturnID"))
            if rid and rid in return_by_id:
                s = n(return_by_id[rid].get("ReturnStatus"))
                if s: rs.append(s)
        exp = derive_aftersales(rs)
        if (fs == "cancelled" and actual != "no_return") or (fs != "cancelled" and actual != exp):
            after_bad += 1
    c.add("RULE_AfterSalesDerived", "RULE", after_bad == 0, "aftersales derived ok", f"mismatches={after_bad}")

    com_wo_score = 0; bad_score = 0; rating_non_del = 0; rating_on_returned = 0
    for ol in orderlines:
        score = n(ol.get("RatingScore")); comment = n(ol.get("Comment")); rid = n(ol.get("ReturnID")); did = n(ol.get("DeliveryID"))
        if score is None and comment is not None: com_wo_score += 1
        if score is not None:
            try:
                s = int(score)
                if s < 0 or s > 10: bad_score += 1
            except ValueError:
                bad_score += 1
            ds = n(delivery_by_id.get(did, {}).get("DeliveryStatus")) if did else None
            if ds != "delivered": rating_non_del += 1
            if rid is not None: rating_on_returned += 1
    c.add("RULE_NoCommentOnly", "RULE", com_wo_score == 0, "no comment-only", f"count={com_wo_score}")
    c.add("RULE_ScoreRange", "RULE", bad_score == 0, "score range valid", f"bad score rows={bad_score}")
    c.add("LIFE_Rating_DeliveredOnly", "LIFECYCLE", rating_non_del == 0, "rating on delivered only", f"violations={rating_non_del}")
    c.add("LIFE_Rating_NoReturned", "LIFECYCLE", rating_on_returned == 0, "no ratings on returned lines", f"violations={rating_on_returned}")

    end_bad = 0
    for o in orders:
        fs = n(o.get("FulfillmentStatus")); endd = n(o.get("EndDate"))
        if fs in {"delivered", "cancelled"} and endd is None: end_bad += 1
        if fs not in {"delivered", "cancelled"} and endd is not None: end_bad += 1
    c.add("RULE_EndDate", "RULE", end_bad == 0, "EndDate by fulfillment valid", f"violations={end_bad}")

    # NULL matrix checks by status and field dependency.
    del_null_bad = 0
    del_null_sample = []
    for row in deliveries:
        did = n(row.get("DeliveryID"))
        ds = n(row.get("DeliveryStatus"))
        sh = n(row.get("ShippedDate"))
        de = n(row.get("DeliveredDate"))
        tr = n(row.get("TrackingNumber"))
        bad = False
        if ds in {"pending", "packed"} and (sh is not None or de is not None or tr is not None):
            bad = True
        elif ds == "shipped" and (sh is None or de is not None or tr is None):
            bad = True
        elif ds == "delivered" and (sh is None or de is None or tr is None):
            bad = True
        if bad:
            del_null_bad += 1
            if len(del_null_sample) < 5:
                del_null_sample.append((did, ds, sh, de, tr))
    c.add(
        "NULLMATRIX_DeliveryStatusFields",
        "NULL_MATRIX",
        del_null_bad == 0,
        f"delivery NULL matrix valid; checked={len(deliveries)}",
        f"violations={del_null_bad} sample={del_null_sample}",
    )

    tracking_format_bad = 0
    tracking_checked = 0
    for row in deliveries:
        tr = n(row.get("TrackingNumber"))
        if tr is None:
            continue
        tracking_checked += 1
        if not any(pat.match(tr) for pat in TRACKING_PATTERNS):
            tracking_format_bad += 1
    c.add(
        "FORMAT_TrackingNumber",
        "FORMAT",
        tracking_format_bad == 0,
        f"tracking number format valid; checked={tracking_checked}",
        f"bad tracking format rows={tracking_format_bad}",
    )

    ol_null_bad = 0
    ol_null_sample = []
    for ol in orderlines:
        olid = n(ol.get("OrderlineID"))
        did = n(ol.get("DeliveryID"))
        rid = n(ol.get("ReturnID"))
        score = n(ol.get("RatingScore"))
        rdt = n(ol.get("RatingCreatedAt"))
        com = n(ol.get("Comment"))
        bad = False
        if did is None and (rid is not None or score is not None or rdt is not None or com is not None):
            bad = True
        if score is None and rdt is not None:
            bad = True
        if score is not None and rdt is None:
            bad = True
        if bad:
            ol_null_bad += 1
            if len(ol_null_sample) < 5:
                ol_null_sample.append((olid, did, rid, score, rdt, com))
    c.add(
        "NULLMATRIX_OrderlineDeps",
        "NULL_MATRIX",
        ol_null_bad == 0,
        f"orderline NULL dependency valid; checked={len(orderlines)}",
        f"violations={ol_null_bad} sample={ol_null_sample}",
    )

    ret_null_bad = 0
    for rr in returns:
        if any(n(rr.get(k)) is None for k in ("ReturnReason", "ReturnQty", "ReturnDate", "ReturnStatus")):
            ret_null_bad += 1
    c.add(
        "NULLMATRIX_ReturnRequired",
        "NULL_MATRIX",
        ret_null_bad == 0,
        f"return required fields not NULL; checked={len(returns)}",
        f"violations={ret_null_bad}",
    )

    # Strict time-chain checks: Start <= Shipped <= Delivered <= Return/Rating <= End.
    bad_date_fmt = 0
    for row in orders:
        if n(row.get("StartDate")) is not None and dt(row.get("StartDate")) is None:
            bad_date_fmt += 1
        if n(row.get("EndDate")) is not None and dt(row.get("EndDate")) is None:
            bad_date_fmt += 1
    for row in deliveries:
        if n(row.get("ShippedDate")) is not None and dt(row.get("ShippedDate")) is None:
            bad_date_fmt += 1
        if n(row.get("DeliveredDate")) is not None and dt(row.get("DeliveredDate")) is None:
            bad_date_fmt += 1
    for row in returns:
        if n(row.get("ReturnDate")) is not None and dt(row.get("ReturnDate")) is None:
            bad_date_fmt += 1
    for row in orderlines:
        if n(row.get("RatingCreatedAt")) is not None and dt(row.get("RatingCreatedAt")) is None:
            bad_date_fmt += 1
    c.add("TIME_DateFormat", "TIME", bad_date_fmt == 0, "date fields parseable", f"bad date fields={bad_date_fmt}")

    order_start = {oid: dt(o.get("StartDate")) for oid, o in order_by_id.items()}
    order_end = {oid: dt(o.get("EndDate")) for oid, o in order_by_id.items()}
    delivery_shipped = {did: dt(d.get("ShippedDate")) for did, d in delivery_by_id.items()}
    delivery_delivered = {did: dt(d.get("DeliveredDate")) for did, d in delivery_by_id.items()}
    return_date_by_id = {rid: dt(r.get("ReturnDate")) for rid, r in return_by_id.items()}

    t_start_end = 0
    t_start_ship = 0
    t_ship_del = 0
    t_del_end = 0
    t_del_return = 0
    t_del_rating = 0
    t_return_end = 0
    t_rating_end = 0

    for oid in [x for x in order_ids if x]:
        st = order_start.get(oid)
        en = order_end.get(oid)
        if st is not None and en is not None and st > en:
            t_start_end += 1

    for did in [x for x in delivery_ids if x]:
        sh = delivery_shipped.get(did)
        de = delivery_delivered.get(did)
        if sh is not None and de is not None and sh > de:
            t_ship_del += 1
        for ol in ol_by_delivery.get(did, []):
            oid = n(ol.get("OrderID"))
            st = order_start.get(oid)
            en = order_end.get(oid)
            if st is not None and sh is not None and st > sh:
                t_start_ship += 1
            if de is not None and en is not None and de > en:
                t_del_end += 1

    for ol in orderlines:
        oid = n(ol.get("OrderID"))
        did = n(ol.get("DeliveryID"))
        rid = n(ol.get("ReturnID"))
        st = order_start.get(oid)
        en = order_end.get(oid)
        de = delivery_delivered.get(did) if did else None
        rating_dt = dt(ol.get("RatingCreatedAt"))
        if rid is not None:
            rd = return_date_by_id.get(rid)
            if de is not None and rd is not None and de > rd:
                t_del_return += 1
            if en is not None and rd is not None and rd > en:
                t_return_end += 1
            if st is not None and rd is not None and st > rd:
                t_start_ship += 1
        if rating_dt is not None:
            if de is not None and de > rating_dt:
                t_del_rating += 1
            if en is not None and rating_dt > en:
                t_rating_end += 1
            if st is not None and st > rating_dt:
                t_start_ship += 1

    c.add("TIME_StartLEEnd", "TIME", t_start_end == 0, "all StartDate <= EndDate", f"violations={t_start_end}")
    c.add("TIME_StartLEShipped", "TIME", t_start_ship == 0, "all StartDate <= Shipped/derived event", f"violations={t_start_ship}")
    c.add("TIME_ShippedLEDelivered", "TIME", t_ship_del == 0, "all ShippedDate <= DeliveredDate", f"violations={t_ship_del}")
    c.add("TIME_DeliveredLEEnd", "TIME", t_del_end == 0, "all DeliveredDate <= EndDate", f"violations={t_del_end}")
    c.add("TIME_DeliveredLEReturn", "TIME", t_del_return == 0, "all DeliveredDate <= ReturnDate", f"violations={t_del_return}")
    c.add("TIME_DeliveredLERating", "TIME", t_del_rating == 0, "all DeliveredDate <= RatingCreatedAt", f"violations={t_del_rating}")
    c.add("TIME_ReturnLEEnd", "TIME", t_return_end == 0, "all ReturnDate <= EndDate", f"violations={t_return_end}")
    c.add("TIME_RatingLEEnd", "TIME", t_rating_end == 0, "all RatingCreatedAt <= EndDate", f"violations={t_rating_end}")

    # Amount derivation consistency from Product + Discount + Delivery shipping.
    discount_by_id = {n(x.get("DiscountID")): x for x in discounts}
    amount_bad_order = 0
    amount_bad_lines = 0
    amount_bad_discount = 0
    amount_bad_total = 0
    amount_bad_loss = 0
    amount_subtotal_sum = Decimal("0")
    amount_shipping_sum = Decimal("0")
    amount_discount_sum = Decimal("0")
    amount_total_sum = Decimal("0")
    for oid, o in order_by_id.items():
        lines = ol_by_order.get(oid, [])
        subtotal = Decimal("0")
        dids = set()
        line_unit_price_cost = []
        for ol in lines:
            q = d(ol.get("Quantity"))
            pid = n(ol.get("ProductID"))
            sid = n(ol.get("SupplierID"))
            unit = product_price_by_id.get(pid)
            cost = sp_cost_by_pair.get((sid, pid))
            if q is None or q <= 0 or unit is None or unit < 0 or cost is None or cost < 0:
                amount_bad_lines += 1
                continue
            subtotal += q_money(q * unit)
            line_unit_price_cost.append((unit, cost))
            did = n(ol.get("DeliveryID"))
            if did is not None:
                dids.add(did)
        subtotal = q_money(subtotal)
        shipping = Decimal("0")
        for did in dids:
            ship = d(delivery_by_id.get(did, {}).get("ShippedGBP"))
            if ship is None or ship < 0:
                amount_bad_order += 1
                continue
            shipping += ship
        shipping = q_money(shipping)

        disc_id = n(o.get("DiscountID"))
        disc_row = discount_by_id.get(disc_id)
        if disc_row is None:
            amount_bad_discount += 1
            continue
        disc_type = n(disc_row.get("DiscountType"))
        disc_method = n(disc_row.get("DiscountMethod"))
        disc_value = d(disc_row.get("DiscountValue"))
        if disc_value is None or disc_value < 0:
            amount_bad_discount += 1
            continue
        if disc_type == "no_discount" and not (disc_method == "fixed" and disc_value == Decimal("0")):
            amount_bad_discount += 1
        if disc_method == "fixed":
            disc_amt = disc_value
            disc_pct = clamp_decimal(disc_value / subtotal if subtotal > 0 else Decimal("0"), Decimal("0"), Decimal("1"))
        elif disc_method == "percentage":
            disc_amt = q_money(subtotal * disc_value / Decimal("100"))
            disc_pct = clamp_decimal(disc_value / Decimal("100"), Decimal("0"), Decimal("1"))
        else:
            amount_bad_discount += 1
            continue
        if disc_amt < 0:
            disc_amt = Decimal("0")
        if disc_amt > subtotal:
            disc_amt = subtotal
        disc_amt = q_money(disc_amt)

        for unit, cost in line_unit_price_cost:
            discounted_unit = q_money(unit * (Decimal("1") - disc_pct))
            if discounted_unit + Decimal("0.000001") < cost:
                amount_bad_loss += 1

        total = q_money(subtotal + shipping - disc_amt)
        if total < 0:
            amount_bad_total += 1
        if subtotal <= 0:
            amount_bad_order += 1

        amount_subtotal_sum += subtotal
        amount_shipping_sum += shipping
        amount_discount_sum += disc_amt
        amount_total_sum += total

    c.add(
        "AMT_LineInputs",
        "AMOUNT",
        amount_bad_lines == 0,
        f"amount line inputs valid; lines={len(orderlines)}",
        f"bad line input rows={amount_bad_lines}",
    )
    c.add(
        "AMT_DiscountLogic",
        "AMOUNT",
        amount_bad_discount == 0,
        "discount dictionary/method/value logic valid",
        f"discount logic violations={amount_bad_discount}",
    )
    c.add(
        "AMT_OrderNonNegative",
        "AMOUNT",
        amount_bad_total == 0,
        f"all derived order totals non-negative; checked={len(orders)}",
        f"negative derived totals={amount_bad_total}",
    )
    c.add(
        "AMT_DiscountNoLoss",
        "AMOUNT",
        amount_bad_loss == 0,
        "discounted unit price never below procurement unit cost",
        f"loss-making line checks={amount_bad_loss}",
    )
    c.add(
        "AMT_OrderAggregateLogic",
        "AMOUNT",
        amount_bad_order == 0,
        "order amount aggregation logic valid",
        f"aggregation violations={amount_bad_order}",
    )

    margin_mean = Decimal("0")
    margin_std = Decimal("0")
    if margin_values:
        margin_mean = sum(margin_values, Decimal("0")) / Decimal(len(margin_values))
        variance = sum((x - margin_mean) * (x - margin_mean) for x in margin_values) / Decimal(len(margin_values))
        margin_std = variance.sqrt() if variance >= 0 else Decimal("0")

    c.add(
        "DIST_GrossMarginMean",
        "DISTRIBUTION",
        Decimal("0.45") <= margin_mean <= Decimal("0.55"),
        f"gross margin mean in range [45%,55%]; actual={margin_mean:.2%}",
        f"gross margin mean out of range [45%,55%]; actual={margin_mean:.2%}",
    )
    c.add(
        "DIST_GrossMarginStd",
        "DISTRIBUTION",
        Decimal("0.05") <= margin_std <= Decimal("0.18"),
        f"gross margin std in range [5%,18%]; actual={margin_std:.2%}",
        f"gross margin std out of range [5%,18%]; actual={margin_std:.2%}",
    )

    flat_by_ol = defaultdict(list)
    for r in flat:
        olid = n(r.get("OrderlineID"))
        if olid: flat_by_ol[olid].append(r)
    miss_flat = sum(1 for olid in [x for x in orderline_ids if x] if olid not in flat_by_ol)

    flat_fk_bad = 0; flat_key_bad = 0
    for r in flat:
        oid = n(r.get("OrderID")); cid = n(r.get("CustomerID")); did = n(r.get("DiscountID")); olid = n(r.get("OrderlineID")); pid = n(r.get("ProductID")); sid = n(r.get("SupplierID")); dlv = n(r.get("DeliveryID")); rid = n(r.get("ReturnID")); iid = n(r.get("InventoryID"))
        if oid not in order_by_id or cid not in customer_ids or did not in discount_ids or pid not in product_ids or sid not in supplier_ids:
            flat_fk_bad += 1; continue
        if dlv is not None and dlv not in delivery_ids: flat_fk_bad += 1; continue
        if rid is not None and rid not in return_ids: flat_fk_bad += 1; continue
        if iid is not None and iid not in inventory_ids: flat_fk_bad += 1; continue
        ol = ol_by_id.get(olid)
        if ol is None: flat_fk_bad += 1; continue
        if n(ol.get("OrderID")) != oid or n(ol.get("ProductID")) != pid or n(ol.get("SupplierID")) != sid or n(ol.get("DeliveryID")) != dlv or n(ol.get("ReturnID")) != rid:
            flat_key_bad += 1; continue
        if iid is not None and n(inv_by_id[iid].get("ProductID")) != pid:
            flat_key_bad += 1
    c.add("FLAT_OrderlineCoverage", "FLAT", miss_flat == 0, "all orderlines in flat", f"missing orderlines={miss_flat}")
    c.add("FLAT_KeyRefs", "FLAT", flat_fk_bad == 0, "flat key refs valid", f"fk errors={flat_fk_bad}")
    c.add("FLAT_KeyConsistency", "FLAT", flat_key_bad == 0, "flat key consistency valid", f"key mismatches={flat_key_bad}")

    flat_order_ids = {n(r.get("OrderID")) for r in flat if n(r.get("OrderID"))}
    flat_customer_ids = {n(r.get("CustomerID")) for r in flat if n(r.get("CustomerID"))}
    flat_product_ids = {n(r.get("ProductID")) for r in flat if n(r.get("ProductID"))}
    flat_supplier_ids = {n(r.get("SupplierID")) for r in flat if n(r.get("SupplierID"))}
    flat_discount_ids = {n(r.get("DiscountID")) for r in flat if n(r.get("DiscountID"))}
    flat_orderline_ids = {n(r.get("OrderlineID")) for r in flat if n(r.get("OrderlineID"))}
    flat_delivery_ids = {n(r.get("DeliveryID")) for r in flat if n(r.get("DeliveryID"))}
    flat_return_ids = {n(r.get("ReturnID")) for r in flat if n(r.get("ReturnID"))}
    flat_inventory_ids = {n(r.get("InventoryID")) for r in flat if n(r.get("InventoryID"))}
    flat_sp_pairs = {
        (n(r.get("SupplierID")), n(r.get("ProductID")))
        for r in flat
        if n(r.get("SupplierID")) and n(r.get("ProductID"))
    }
    sp_pairs = {
        (n(r.get("SupplierID")), n(r.get("ProductID")))
        for r in supplier_products
        if n(r.get("SupplierID")) and n(r.get("ProductID"))
    }

    # Only verify inventory rows that were actually involved in a delivery appear in
    # FlatView. Inventory at warehouses with no deliveries legitimately won't appear.
    delivered_pid_wh: set = set()
    for ol in orderlines:
        did_v = n(ol.get("DeliveryID"))
        pid_v = n(ol.get("ProductID"))
        if did_v and pid_v:
            wh_v = n(delivery_by_id.get(did_v, {}).get("Warehouse"))
            if wh_v:
                delivered_pid_wh.add((pid_v, wh_v))
    expected_inv_ids: set = {
        n(inv.get("InventoryID"))
        for inv in inventory
        if (n(inv.get("ProductID")), n(inv.get("Location"))) in delivered_pid_wh
        and n(inv.get("InventoryID"))
    }
    inv_miss = sum(1 for iid in expected_inv_ids if iid not in flat_inventory_ids)

    # For FlatView rows that have a Warehouse value, Location must equal Warehouse
    # and InventoryID must be non-null.
    flat_del_inv_bad = 0
    flat_del_inv_sample = []
    for r in flat:
        wh_r = n(r.get("Warehouse"))
        loc_r = n(r.get("Location"))
        iid_r = n(r.get("InventoryID"))
        if wh_r is not None:
            if iid_r is None or loc_r != wh_r:
                flat_del_inv_bad += 1
                if len(flat_del_inv_sample) < 5:
                    flat_del_inv_sample.append((n(r.get("OrderlineID")), wh_r, loc_r, iid_r))

    c.add(
        "FLAT_InventoryCoverage",
        "FLAT",
        inv_miss == 0,
        f"all delivered-warehouse inventory ids in flat; checked={len(expected_inv_ids)}",
        f"missing inventory ids={inv_miss}",
    )
    c.add(
        "FLAT_DeliveredInventoryConsistency",
        "FLAT",
        flat_del_inv_bad == 0,
        "flat delivered rows: Location == Warehouse and InventoryID non-null",
        f"violations={flat_del_inv_bad} sample={flat_del_inv_sample}",
    )

    miss_order_key = sum(1 for x in order_ids if x and x not in flat_order_ids)
    miss_customer_key = sum(1 for x in customer_ids if x and x not in flat_customer_ids)
    miss_product_key = sum(1 for x in product_ids if x and x not in flat_product_ids)
    miss_supplier_key = sum(1 for x in supplier_ids if x and x not in flat_supplier_ids)
    miss_discount_key = sum(1 for x in discount_ids if x and x not in flat_discount_ids)
    miss_orderline_key = sum(1 for x in orderline_ids if x and x not in flat_orderline_ids)
    miss_delivery_key = sum(1 for x in delivery_ids if x and x not in flat_delivery_ids)
    miss_return_key = sum(1 for x in return_ids if x and x not in flat_return_ids)
    miss_inventory_key = sum(1 for x in expected_inv_ids if x and x not in flat_inventory_ids)
    miss_sp_pair = sum(1 for p in sp_pairs if p not in flat_sp_pairs)

    c.add("FLAT_COVER_OrderID", "FLAT_COVERAGE", miss_order_key == 0, "all OrderID covered in flat", f"missing OrderID count={miss_order_key}")
    c.add("FLAT_COVER_CustomerID", "FLAT_COVERAGE", miss_customer_key == 0, "all CustomerID covered in flat", f"missing CustomerID count={miss_customer_key}")
    c.add("FLAT_COVER_ProductID", "FLAT_COVERAGE", miss_product_key == 0, "all ProductID covered in flat", f"missing ProductID count={miss_product_key}")
    c.add("FLAT_COVER_SupplierID", "FLAT_COVERAGE", miss_supplier_key == 0, "all SupplierID covered in flat", f"missing SupplierID count={miss_supplier_key}")
    c.add("FLAT_COVER_DiscountID", "FLAT_COVERAGE", miss_discount_key == 0, "all DiscountID covered in flat", f"missing DiscountID count={miss_discount_key}")
    c.add("FLAT_COVER_OrderlineID", "FLAT_COVERAGE", miss_orderline_key == 0, "all OrderlineID covered in flat", f"missing OrderlineID count={miss_orderline_key}")
    c.add("FLAT_COVER_DeliveryID", "FLAT_COVERAGE", miss_delivery_key == 0, "all DeliveryID covered in flat", f"missing DeliveryID count={miss_delivery_key}")
    c.add("FLAT_COVER_ReturnID", "FLAT_COVERAGE", miss_return_key == 0, "all ReturnID covered in flat", f"missing ReturnID count={miss_return_key}")
    c.add("FLAT_COVER_InventoryID", "FLAT_COVERAGE", miss_inventory_key == 0, "all delivered-warehouse InventoryID covered in flat", f"missing InventoryID count={miss_inventory_key}")
    c.add("FLAT_COVER_SupplierProduct", "FLAT_COVERAGE", miss_sp_pair == 0, "all SupplierProduct pairs covered in flat", f"missing SupplierProduct pair count={miss_sp_pair}")

    # Distribution checks
    fcnt = Counter(n(x.get("FulfillmentStatus")) for x in orders)
    pcnt = Counter(n(x.get("Platform")) for x in orders)
    delivered_lines = [ol for ol in orderlines if n(delivery_by_id.get(n(ol.get("DeliveryID")), {}).get("DeliveryStatus")) == "delivered"]
    completed_lines = [ol for ol in delivered_lines if n(ol.get("ReturnID")) is None]
    rated = sum(1 for ol in orderlines if n(ol.get("RatingScore")) is not None)
    commented = sum(1 for ol in orderlines if n(ol.get("Comment")) is not None)
    commented_completed = sum(1 for ol in completed_lines if n(ol.get("Comment")) is not None)
    returned_lines = sum(1 for ol in orderlines if n(ol.get("ReturnID")) is not None)

    delivered_ratio = dec_ratio(fcnt.get("delivered", 0), len(orders))
    comment_ratio_completed = dec_ratio(commented_completed, len(completed_lines))
    return_rate = dec_ratio(returned_lines, len(orderlines))

    c.add(
        "DIST_DeliveredRatio",
        "DISTRIBUTION",
        DELIVERED_RATIO_MIN <= delivered_ratio <= DELIVERED_RATIO_MAX,
        (
            f"delivered ratio in range; target={TARGET_DELIVERED_RATIO:.2%}, "
            f"range=[{DELIVERED_RATIO_MIN:.2%},{DELIVERED_RATIO_MAX:.2%}], actual={delivered_ratio:.2%}"
        ),
        (
            f"delivered ratio out of range; target={TARGET_DELIVERED_RATIO:.2%}, "
            f"range=[{DELIVERED_RATIO_MIN:.2%},{DELIVERED_RATIO_MAX:.2%}], actual={delivered_ratio:.2%}"
        ),
    )
    c.add(
        "DIST_CommentRateCompleted",
        "DISTRIBUTION",
        COMMENT_RATIO_MIN <= comment_ratio_completed <= COMMENT_RATIO_MAX,
        (
            f"comment rate (completed lines) in range; target={TARGET_COMMENT_RATIO_COMPLETED:.2%}, "
            f"range=[{COMMENT_RATIO_MIN:.2%},{COMMENT_RATIO_MAX:.2%}], actual={comment_ratio_completed:.2%}"
        ),
        (
            f"comment rate (completed lines) out of range; target={TARGET_COMMENT_RATIO_COMPLETED:.2%}, "
            f"range=[{COMMENT_RATIO_MIN:.2%},{COMMENT_RATIO_MAX:.2%}], actual={comment_ratio_completed:.2%}"
        ),
    )
    c.add(
        "DIST_ReturnRate",
        "DISTRIBUTION",
        RETURN_RATE_MIN <= return_rate <= RETURN_RATE_MAX,
        (
            f"return rate in range; target={TARGET_RETURN_RATE:.2%}, "
            f"range=[{RETURN_RATE_MIN:.2%},{RETURN_RATE_MAX:.2%}], actual={return_rate:.2%}"
        ),
        (
            f"return rate out of range; target={TARGET_RETURN_RATE:.2%}, "
            f"range=[{RETURN_RATE_MIN:.2%},{RETURN_RATE_MAX:.2%}], actual={return_rate:.2%}"
        ),
    )

    valid_scores = []
    for ol in orderlines:
        s = n(ol.get("RatingScore"))
        if s is None:
            continue
        try:
            sv = int(s)
        except ValueError:
            continue
        if 0 <= sv <= 10:
            valid_scores.append(sv)
    score_count = len(valid_scores)
    score_band_counts = Counter(score_band(x) for x in valid_scores)
    for b in ("0", "1-3", "4-6", "7-8", "9-10"):
        actual = dec_ratio(score_band_counts.get(b, 0), score_count)
        target = TARGET_SCORE_BANDS[b]
        low = max(Decimal("0"), target - SCORE_BAND_TOL)
        high = min(Decimal("1"), target + SCORE_BAND_TOL)
        c.add(
            f"DIST_ScoreBand_{b.replace('-', '_')}",
            "DISTRIBUTION",
            low <= actual <= high,
            f"score band {b} in range; target={target:.2%}, range=[{low:.2%},{high:.2%}], actual={actual:.2%}, n={score_count}",
            f"score band {b} out of range; target={target:.2%}, range=[{low:.2%},{high:.2%}], actual={actual:.2%}, n={score_count}",
        )

    # ID gap checks
    id_values = {
        "OrderID": order_ids,
        "CustomerID": customer_ids,
        "ProductID": product_ids,
        "OrderlineID": orderline_ids,
        "SupplierID": supplier_ids,
        "DiscountID": discount_ids,
        "InventoryID": inventory_ids,
        "DeliveryID": delivery_ids,
        "ReturnID": return_ids,
    }
    gap_stats = {}
    for key, vals in id_values.items():
        events, bad_size, rate, max_gap, total_missing = id_gap_stats({x for x in vals if x})
        gap_stats[key] = (events, bad_size, rate, max_gap, total_missing)
        if key in GAP_CHECK_ID_KEYS:
            c.add(
                f"GAP_{key}_Rate",
                "ID_GAP",
                rate <= ID_GAP_MAX_RATE,
                f"{key} gap-event rate valid; max={ID_GAP_MAX_RATE:.2%}, actual={rate:.2%}, events={events}",
                f"{key} gap-event rate too high; max={ID_GAP_MAX_RATE:.2%}, actual={rate:.2%}, events={events}",
            )
            c.add(
                f"GAP_{key}_Size",
                "ID_GAP",
                bad_size == 0,
                f"{key} gap-size range valid [{ID_GAP_MIN_SIZE},{ID_GAP_MAX_SIZE}]",
                f"{key} gap-size out of range count={bad_size}, max_gap={max_gap}",
            )
        if key in NO_GAP_ID_KEYS:
            c.add(
                f"GAP_{key}_NoGap",
                "ID_GAP",
                events == 0,
                f"{key} has no gap as required; events=0",
                f"{key} unexpected gaps found; events={events}, rate={rate:.2%}, max_gap={max_gap}",
            )

    metrics = {
        "orders": len(orders), "orderlines": len(orderlines), "flat_rows": len(flat),
        "orderline_per_order": round(len(orderlines) / max(1, len(orders)), 3),
        "fulfilled_delivered_ratio": round(float(delivered_ratio), 6),
        "rating_coverage_on_delivered": round(rated / max(1, len(delivered_lines)), 4),
        "comment_rate_on_delivered": round(commented / max(1, len(delivered_lines)), 4),
        "comment_rate_on_completed_lines": round(float(comment_ratio_completed), 6),
        "return_rate_on_orderline": round(float(return_rate), 6),
        "target_delivered_ratio": float(TARGET_DELIVERED_RATIO),
        "target_comment_ratio_completed": float(TARGET_COMMENT_RATIO_COMPLETED),
        "target_return_rate": float(TARGET_RETURN_RATE),
        "target_delivered_ratio_min": float(DELIVERED_RATIO_MIN),
        "target_delivered_ratio_max": float(DELIVERED_RATIO_MAX),
        "target_comment_ratio_min": float(COMMENT_RATIO_MIN),
        "target_comment_ratio_max": float(COMMENT_RATIO_MAX),
        "target_return_rate_min": float(RETURN_RATE_MIN),
        "target_return_rate_max": float(RETURN_RATE_MAX),
        "amount_subtotal_sum": str(q_money(amount_subtotal_sum)),
        "amount_shipping_sum": str(q_money(amount_shipping_sum)),
        "amount_discount_sum": str(q_money(amount_discount_sum)),
        "amount_total_sum": str(q_money(amount_total_sum)),
        "flat_missing_order_keys": miss_order_key,
        "flat_missing_customer_keys": miss_customer_key,
        "flat_missing_product_keys": miss_product_key,
        "flat_missing_supplier_keys": miss_supplier_key,
        "flat_missing_discount_keys": miss_discount_key,
        "flat_missing_orderline_keys": miss_orderline_key,
        "flat_missing_delivery_keys": miss_delivery_key,
        "flat_missing_return_keys": miss_return_key,
        "flat_missing_inventory_keys": miss_inventory_key,
        "flat_missing_supplierproduct_pairs": miss_sp_pair,
        "time_viol_start_end": t_start_end,
        "time_viol_start_ship": t_start_ship,
        "time_viol_ship_del": t_ship_del,
        "time_viol_del_end": t_del_end,
        "time_viol_del_return": t_del_return,
        "time_viol_del_rating": t_del_rating,
        "time_viol_return_end": t_return_end,
        "time_viol_rating_end": t_rating_end,
        "nullmatrix_delivery_viol": del_null_bad,
        "tracking_checked": tracking_checked,
        "tracking_format_bad": tracking_format_bad,
        "nullmatrix_orderline_viol": ol_null_bad,
        "nullmatrix_return_viol": ret_null_bad,
        "amount_bad_lines": amount_bad_lines,
        "amount_bad_order_agg": amount_bad_order,
        "amount_bad_discount": amount_bad_discount,
        "amount_bad_negative_total": amount_bad_total,
        "amount_bad_lossmaking_lines": amount_bad_loss,
        "sp_bad_leadtime": sp_bad_leadtime,
        "sp_bad_unitcost": sp_bad_cost,
        "sp_margin_out_of_range": margin_out_of_range,
        "sp_leadtime_cost_trend_inversions": trend_inversions,
        "sp_leadtime_cost_comparable_pairs": trend_pairs,
        "sp_gross_margin_mean": round(float(margin_mean), 6),
        "sp_gross_margin_std": round(float(margin_std), 6),
    }
    for k, v in sorted(fcnt.items()): metrics[f"fulfillment_{k}"] = v
    for k, v in sorted(pcnt.items()): metrics[f"platform_{k}"] = v
    for b in ("0", "1-3", "4-6", "7-8", "9-10"):
        cnt = score_band_counts.get(b, 0)
        metrics[f"scoreband_count_{b.replace('-', '_')}"] = cnt
        metrics[f"scoreband_ratio_{b.replace('-', '_')}"] = round(float(dec_ratio(cnt, score_count)), 6)
        metrics[f"scoreband_target_{b.replace('-', '_')}"] = float(TARGET_SCORE_BANDS[b])
    for key, (events, bad_size, rate, max_gap, total_missing) in gap_stats.items():
        metrics[f"gap_events_{key}"] = events
        metrics[f"gap_rate_{key}"] = round(float(rate), 6)
        metrics[f"gap_bad_size_{key}"] = bad_size
        metrics[f"gap_max_{key}"] = max_gap
        metrics[f"gap_total_missing_{key}"] = total_missing

    return metrics


def write_report(path: Path, checks, metrics, input_path: Path):
    wb = Workbook()
    ws = wb.active; ws.title = "summary"
    p = sum(1 for x in checks if x.status == "PASS")
    f = sum(1 for x in checks if x.status == "FAIL")
    w = sum(1 for x in checks if x.status == "WARN")
    for k, v in [
        ("generated_at", datetime.now().isoformat(timespec="seconds")),
        ("input_workbook", str(input_path.resolve())),
        ("total_checks", len(checks)), ("pass", p), ("fail", f), ("warn", w),
    ]:
        ws.append([k, v])

    wc = wb.create_sheet("checks"); wc.append(["check_id", "category", "status", "details"])
    for r in checks: wc.append([r.check_id, r.category, r.status, r.details])

    wf = wb.create_sheet("check_fail_warn")
    wf.append(["check_id", "category", "status", "details"])
    for r in checks:
        if r.status in {"FAIL", "WARN"}:
            wf.append([r.check_id, r.category, r.status, r.details])

    by_cat_status = defaultdict(int)
    for r in checks:
        by_cat_status[(r.category, r.status)] += 1
    wcs = wb.create_sheet("check_category_stats")
    wcs.append(["category", "status", "count"])
    for (cat, st), cnt in sorted(by_cat_status.items()):
        wcs.append([cat, st, cnt])

    wm = wb.create_sheet("metrics"); wm.append(["metric", "value"])
    for k in sorted(metrics.keys()): wm.append([k, metrics[k]])
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="Validate generated BiroCo 3NF workbook.")
    ap.add_argument("--input", default=DEFAULT_INPUT_XLSX)
    ap.add_argument("--report", default=DEFAULT_REPORT_XLSX)
    args = ap.parse_args()

    inp = _res(args.input); rep = _res(args.report)
    if not inp.exists():
        raise FileNotFoundError(f"Input workbook not found: {inp}")

    c = Collector()
    t = load_tables(inp)
    missing = [s for s in REQUIRED_SHEETS if s not in t]
    c.add("FILES_RequiredSheets", "FILES", len(missing) == 0, "all sheets present", f"missing sheets: {missing}")
    if missing:
        write_report(rep, c.rows, {}, inp)
        print("Validation failed due to missing sheets.")
        print(f"- report: {rep.resolve()}")
        raise SystemExit(1)

    metrics = validate(c, t)
    write_report(rep, c.rows, metrics, inp)

    # --- per-check results grouped by category ---
    rows_sorted = sorted(c.rows, key=lambda r: r.category)
    print("\nValidation results:")
    for cat, group in groupby(rows_sorted, key=lambda r: r.category):
        print(f"\n  [{cat}]")
        for row in group:
            icon = {"PASS": "+", "FAIL": "x", "WARN": "!"}.get(row.status, "?")
            print(f"    {icon} {row.check_id}: {row.details}")

    # --- summary ---
    n_pass = sum(1 for x in c.rows if x.status == "PASS")
    n_fail = sum(1 for x in c.rows if x.status == "FAIL")
    n_warn = sum(1 for x in c.rows if x.status == "WARN")
    total  = len(c.rows)
    print(f"\n  ── Summary ({total} checks) ──")
    print(f"    PASS  {n_pass:>4}")
    print(f"    WARN  {n_warn:>4}")
    print(f"    FAIL  {n_fail:>4}")

    # --- metrics ---
    if metrics:
        print("\n  ── Metrics ──")
        for k in sorted(metrics.keys()):
            print(f"    {k} = {metrics[k]}")

    print(f"\n  Report: {rep.resolve()}")

    if n_fail > 0:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
